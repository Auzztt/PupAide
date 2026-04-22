import sys
import os
import hashlib
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtGui import QFont
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import subprocess
from PyQt5.QtGui import QBrush
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QStackedWidget, QListWidget,
                             QListWidgetItem, QLabel, QFrame, QMessageBox, QFileDialog,
                             QTextEdit, QProgressBar, QGroupBox, QCheckBox,
                             QLineEdit, QSplitter, QTabWidget, QComboBox, QFormLayout,
                             QSlider, QColorDialog, QAbstractItemView, QSpinBox,
                             QTreeWidget, QTreeWidgetItem, QHeaderView, QTableWidget, QTableWidgetItem)
from PyQt5.QtGui import QColor
import PyPDF2
import pdfplumber
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import io


# 获取全局应用对象，用于后续强制刷新样式
qApp = None

# TODO-1 添加全局查找文件的功能，并添加到菜单栏中


class PupAideMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("小狗助理 PupAide - 办公百宝箱")
        self.setMinimumSize(1200, 800)

        # 设置主窗口样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
            /* 为 QMessageBox 中的按钮单独设置样式 */
            QMessageBox QPushButton {
                background-color: #f0f0f0;
                color: #333;
                border: 1px solid #ccc;
                padding: 6px 12px;
                min-height: 28px;
            }
            QMessageBox QPushButton:hover {
                background-color: #e0e0e0;
            }

            QListWidget {
                border: 1px solid #ddd;
                border-radius: 6px;
                background-color: white;
            }
            QListWidget::item {
                padding: 10px 8px;
            }
            QListWidget::item:selected {
                background-color: #FFB347;
                color: white;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QTextEdit {
                border: 1px solid #ddd;
                border-radius: 6px;
                font-family: monospace;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 6px;
                min-height: 30px;
            }
            /* 关键：只调整勾选框大小，保持原生对勾样式 */
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
            }
            QProgressBar {
                border: 1px solid #ddd;
                border-radius: 6px;
                text-align: center;
                min-height: 25px;
            }
            QProgressBar::chunk {
                background-color: #FFB347;
                border-radius: 5px;
            }
            QSplitter::handle {
                background-color: #ddd;
                width: 4px;
            }
            QSplitter::handle:hover {
                background-color: #FFB347;
            }
        """)

        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(0)

        # ========== 使用QSplitter实现可拖拽调整大小 ==========
        self.splitter = QSplitter(Qt.Horizontal)

        # ========== 左侧面板 ==========
        left_panel = QFrame()
        left_panel.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 8px;
                border: 1px solid #ddd;
            }
        """)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(15, 20, 15, 20)

        # 标题
        title_label = QLabel("🐕 小狗助理")
        title_font = QFont()
        title_font.setPointSize(18)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #FFB347; margin-bottom: 20px;")
        title_label.setWordWrap(True)
        left_layout.addWidget(title_label)

        # 功能列表
        self.nav_list = QListWidget()
        list_font = QFont()
        list_font.setPointSize(14)
        self.nav_list.setFont(list_font)
        functions = [
            "📁 重复文件查找器",
            "🔄 文件夹同步/备份",
            "💾 磁盘空间分析器",
            "📄 PDF 批量处理",
            "📝 Word/Excel 批量替换",
            "🔍 OCR 文字识别工具",
            "✂️ 文件名称提取器"  # 新增功能
        ]
        for func in functions:
            item = QListWidgetItem(func)
            item.setSizeHint(QSize(0, 45))
            self.nav_list.addItem(item)
        left_layout.addWidget(self.nav_list)

        # 底部信息
        info_label = QLabel("版本 1.1.0\n拖动右侧边缘可调整菜单宽度")
        info_font = QFont()
        info_font.setPointSize(12)
        info_label.setFont(info_font)
        info_label.setAlignment(Qt.AlignCenter)
        info_label.setStyleSheet("color: #999; margin-top: 15px;")
        info_label.setWordWrap(True)
        left_layout.addWidget(info_label)
        left_layout.addStretch()

        # ========== 右侧内容区域 ==========
        self.right_panel = QStackedWidget()
        self.right_panel.setStyleSheet("""
            QWidget {
                background-color: white;
                border-radius: 8px;
                border: 1px solid #ddd;
            }
        """)

        # 创建各个功能页面
        self.page_duplicate = DuplicateFilePage()
        self.page_sync = SyncBackupPage()
        self.page_disk = DiskAnalyzerPage()
        self.page_pdf = PDFBatchPage()
        self.page_office = OfficeBatchPage()
        self.page_ocr = OCRPage()
        self.page_namecut = FileNameCut()  # 新增功能页面实例

        self.right_panel.addWidget(self.page_duplicate)
        self.right_panel.addWidget(self.page_sync)
        self.right_panel.addWidget(self.page_disk)
        self.right_panel.addWidget(self.page_pdf)
        self.right_panel.addWidget(self.page_office)
        self.right_panel.addWidget(self.page_ocr)
        self.right_panel.addWidget(self.page_namecut)  # 新增功能页面添加到右侧面板

        # 添加导航栏点击事件
        self.nav_list.currentRowChanged.connect(
            self.right_panel.setCurrentIndex)

        # 将左右面板添加到分割器
        self.splitter.addWidget(left_panel)
        self.splitter.addWidget(self.right_panel)

        # 设置初始大小比例
        self.splitter.setSizes([250, 800])

        # 添加到主布局
        main_layout.addWidget(self.splitter)


# ========== 功能1：重复文件查找器 ==========
class DuplicateFilePage(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 标题
        title = QLabel("📁 重复文件查找器")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # 文件夹选择区域
        folder_group = QGroupBox("选择要扫描的文件夹")
        group_font = QFont()
        group_font.setPointSize(13)
        folder_group.setFont(group_font)
        folder_layout = QVBoxLayout(folder_group)
        folder_layout.setSpacing(10)

        # 文件夹路径选择行
        path_layout = QHBoxLayout()
        self.folder_path = QLineEdit()
        path_font = QFont()
        path_font.setPointSize(13)
        self.folder_path.setFont(path_font)
        self.folder_path.setPlaceholderText("请选择文件夹...")
        self.folder_path.setReadOnly(True)

        # 选择文件夹按钮
        btn_select = QPushButton("📂 选择文件夹")
        btn_select.setMinimumWidth(130)
        btn_select.clicked.connect(self.select_folder)

        # 确保文本显示
        btn_select.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        path_layout.addWidget(self.folder_path)
        path_layout.addWidget(btn_select)
        folder_layout.addLayout(path_layout)

        layout.addWidget(folder_group)

        # 扫描选项
        option_group = QGroupBox("扫描选项")
        option_group.setFont(group_font)
        option_layout = QVBoxLayout(option_group)
        option_layout.setSpacing(10)

        # 勾选框：只调整大小，不改变勾选样式
        self.check_subfolders = QCheckBox("包含子文件夹")
        check_font = QFont()
        check_font.setPointSize(10)
        self.check_subfolders.setFont(check_font)
        self.check_subfolders.setChecked(True)
        # 强制只调整大小，不影响默认勾选图标
        self.check_subfolders.setStyleSheet("""
            QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
            }
        """)

        option_layout.addWidget(self.check_subfolders)
        layout.addWidget(option_group)

        # 按钮行
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        # 扫描按钮
        self.btn_scan = QPushButton("🔍 开始扫描")
        self.btn_scan.setMinimumWidth(130)
        self.btn_scan.clicked.connect(self.scan_files)
        self.btn_scan.setEnabled(False)

        # 确保文本显示
        self.btn_scan.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        # 清空按钮
        self.btn_clear = QPushButton("🗑️ 清空结果")
        self.btn_clear.setMinimumWidth(120)
        self.btn_clear.clicked.connect(self.clear_results)

        # 确保文本显示
        self.btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        button_layout.addWidget(self.btn_scan)
        button_layout.addWidget(self.btn_clear)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # 结果显示区域
        result_group = QGroupBox("扫描结果")
        result_group.setFont(group_font)
        result_layout = QVBoxLayout(result_group)

        self.result_text = QTextEdit()
        text_font = QFont()
        text_font.setPointSize(12)
        self.result_text.setFont(text_font)
        self.result_text.setReadOnly(True)
        self.result_text.setMinimumHeight(350)
        result_layout.addWidget(self.result_text)

        layout.addWidget(result_group)

        self.scan_folder = None

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择文件夹")
        if folder:
            self.scan_folder = folder
            self.folder_path.setText(folder)
            self.btn_scan.setEnabled(True)
            self.result_text.clear()
            self.result_text.append(f"✅ 已选择文件夹: {folder}\n")
            self.result_text.append("点击「开始扫描」查找重复文件...\n")

    def clear_results(self):
        self.result_text.clear()
        if self.scan_folder:
            self.result_text.append(f"✅ 已选择文件夹: {self.scan_folder}\n")
            self.result_text.append("点击「开始扫描」查找重复文件...\n")

    def scan_files(self):
        if not self.scan_folder:
            QMessageBox.warning(self, "提示", "请先选择文件夹！")
            return

        self.btn_scan.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.result_text.clear()
        self.result_text.append(f"🔍 正在扫描文件夹: {self.scan_folder}\n")
        self.result_text.append("=" * 60 + "\n\n")

        self.progress.setValue(30)

        file_dict = {}
        total_files = 0

        # 遍历文件
        if self.check_subfolders.isChecked():
            for root, dirs, files in os.walk(self.scan_folder):
                for file in files:
                    total_files += 1
        else:
            for file in os.listdir(self.scan_folder):
                if os.path.isfile(os.path.join(self.scan_folder, file)):
                    total_files += 1

        if total_files == 0:
            self.result_text.append("❌ 没有找到任何文件\n")
            self.progress.setVisible(False)
            self.btn_scan.setEnabled(True)
            return

        self.progress.setValue(50)
        self.result_text.append(f"📊 找到 {total_files} 个文件，正在分析...\n\n")

        # 计算文件哈希
        scanned = 0
        if self.check_subfolders.isChecked():
            for root, dirs, files in os.walk(self.scan_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    try:
                        if os.path.getsize(file_path) > 1024:
                            with open(file_path, 'rb') as f:
                                file_hash = hashlib.md5(
                                    f.read(8192)).hexdigest()

                            key = (file, os.path.getsize(file_path))
                            if key not in file_dict:
                                file_dict[key] = []
                            file_dict[key].append(file_path)
                    except:
                        pass
                    scanned += 1
                    if scanned % 10 == 0:
                        self.progress.setValue(
                            50 + int(scanned / total_files * 40))
        else:
            for file in os.listdir(self.scan_folder):
                file_path = os.path.join(self.scan_folder, file)
                if os.path.isfile(file_path):
                    try:
                        if os.path.getsize(file_path) > 1024:
                            with open(file_path, 'rb') as f:
                                file_hash = hashlib.md5(
                                    f.read(8192)).hexdigest()

                            key = (file, os.path.getsize(file_path))
                            if key not in file_dict:
                                file_dict[key] = []
                            file_dict[key].append(file_path)
                    except:
                        pass
                    scanned += 1
                    if scanned % 10 == 0:
                        self.progress.setValue(
                            50 + int(scanned / total_files * 40))

        self.progress.setValue(100)

        # 显示结果
        duplicates_found = False
        for (filename, size), paths in file_dict.items():
            if len(paths) > 1:
                duplicates_found = True
                size_kb = size / 1024
                if size_kb > 1024:
                    size_str = f"{size_kb/1024:.2f} MB"
                else:
                    size_str = f"{size_kb:.2f} KB"

                self.result_text.append(f"📄 {filename} ({size_str})\n")
                for path in paths:
                    self.result_text.append(f"   └─ {path}\n")
                self.result_text.append("\n")

        if not duplicates_found:
            self.result_text.append("✅ 恭喜！没有找到重复文件！\n")
        else:
            self.result_text.append(
                f"\n✅ 扫描完成！共找到 {len([k for k,v in file_dict.items() if len(v)>1])} 组重复文件\n")

        self.progress.setVisible(False)
        self.btn_scan.setEnabled(True)


# ========== 功能2：文件夹同步/备份工具 ==========
class SyncBackupPage(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 标题
        title = QLabel("🔄 文件夹同步/备份工具")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # 源文件夹选择
        source_group = QGroupBox("源文件夹（要备份的文件夹）")
        group_font = QFont()
        group_font.setPointSize(13)
        source_group.setFont(group_font)
        source_layout = QVBoxLayout(source_group)
        source_layout.setSpacing(10)

        source_path_layout = QHBoxLayout()
        self.source_path = QLineEdit()
        self.source_path.setPlaceholderText("请选择要备份的源文件夹...")
        self.source_path.setReadOnly(True)

        self.btn_source = QPushButton("📂 选择源文件夹")
        self.btn_source.setMinimumWidth(130)
        self.btn_source.clicked.connect(self.select_source_folder)
        self.btn_source.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        source_path_layout.addWidget(self.source_path)
        source_path_layout.addWidget(self.btn_source)
        source_layout.addLayout(source_path_layout)
        layout.addWidget(source_group)

        # 目标文件夹选择
        target_group = QGroupBox("目标文件夹（备份保存位置）")
        target_group.setFont(group_font)
        target_layout = QVBoxLayout(target_group)
        target_layout.setSpacing(10)

        target_path_layout = QHBoxLayout()
        self.target_path = QLineEdit()
        self.target_path.setPlaceholderText("请选择备份保存的目标文件夹...")
        self.target_path.setReadOnly(True)

        self.btn_target = QPushButton("📂 选择目标文件夹")
        self.btn_target.setMinimumWidth(130)
        self.btn_target.clicked.connect(self.select_target_folder)
        self.btn_target.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        target_path_layout.addWidget(self.target_path)
        target_path_layout.addWidget(self.btn_target)
        target_layout.addLayout(target_path_layout)
        layout.addWidget(target_group)

        # 同步选项
        option_group = QGroupBox("同步选项")
        option_group.setFont(group_font)
        option_layout = QVBoxLayout(option_group)

        self.check_subfolders = QCheckBox("包含子文件夹")
        check_font = QFont()
        check_font.setPointSize(10)
        self.check_subfolders.setFont(check_font)
        self.check_subfolders.setChecked(True)
        self.check_subfolders.setStyleSheet("""
            QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
            }
        """)

        self.check_overwrite = QCheckBox("覆盖已存在的文件")
        check_font = QFont()
        check_font.setPointSize(10)
        self.check_overwrite.setFont(check_font)
        self.check_overwrite.setChecked(True)
        self.check_overwrite.setStyleSheet("""
            QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
            }
        """)

        self.check_delete = QCheckBox("删除目标文件夹中多余的文件（保持完全一致）")
        check_font = QFont()
        check_font.setPointSize(10)
        self.check_delete.setFont(check_font)
        self.check_delete.setChecked(False)
        self.check_delete.setStyleSheet("""
                    QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
            }
        """)

        option_layout.addWidget(self.check_subfolders)
        option_layout.addWidget(self.check_overwrite)
        option_layout.addWidget(self.check_delete)
        layout.addWidget(option_group)

        # 按钮行
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        self.btn_sync = QPushButton("🔄 开始同步")
        self.btn_sync.setMinimumWidth(130)
        self.btn_sync.clicked.connect(self.start_sync)
        self.btn_sync.setEnabled(False)
        self.btn_sync.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        self.btn_preview = QPushButton("👁️ 预览同步")
        self.btn_preview.setMinimumWidth(120)
        self.btn_preview.clicked.connect(self.preview_sync)
        self.btn_preview.setEnabled(False)
        self.btn_preview.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        button_layout.addWidget(self.btn_sync)
        button_layout.addWidget(self.btn_preview)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # 结果显示区域
        result_group = QGroupBox("同步日志")
        result_group.setFont(group_font)
        result_layout = QVBoxLayout(result_group)

        self.result_text = QTextEdit()
        text_font = QFont()
        text_font.setPointSize(12)
        self.result_text.setFont(text_font)
        self.result_text.setReadOnly(True)
        self.result_text.setMinimumHeight(350)
        result_layout.addWidget(self.result_text)

        layout.addWidget(result_group)

        self.source_folder = None
        self.target_folder = None

    def select_source_folder(self):
        try:
            # 使用文件对话框选择文件夹
            folder = QFileDialog.getExistingDirectory(
                self,
                "选择源文件夹"
            )
            if folder and os.path.exists(folder):
                self.source_folder = folder
                self.source_path.setText(folder)
                self.check_buttons_enabled()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"选择文件夹失败：{str(e)}")

    def select_target_folder(self):
        try:
            folder = QFileDialog.getExistingDirectory(
                self,
                "选择目标文件夹"
            )
            if folder and os.path.exists(folder):
                self.target_folder = folder
                self.target_path.setText(folder)
                self.check_buttons_enabled()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"选择文件夹失败：{str(e)}")

    def check_buttons_enabled(self):
        enabled = bool(self.source_folder and self.target_folder)
        self.btn_sync.setEnabled(enabled)
        self.btn_preview.setEnabled(enabled)

    def preview_sync(self):
        if not self.source_folder or not self.target_folder:
            QMessageBox.warning(self, "提示", "请先选择源文件夹和目标文件夹！")
            return

        self.result_text.clear()
        self.result_text.append("🔍 正在分析同步内容...\n")
        self.result_text.append("=" * 60 + "\n\n")

        # 获取需要同步的文件列表
        files_to_copy = []
        files_to_delete = []

        # 收集源文件夹中的所有文件
        source_files = {}
        try:
            if self.check_subfolders.isChecked():
                for root, dirs, files in os.walk(self.source_folder):
                    for file in files:
                        full_path = os.path.join(root, file)
                        rel_path = os.path.relpath(
                            full_path, self.source_folder)
                        source_files[rel_path] = full_path
            else:
                for item in os.listdir(self.source_folder):
                    full_path = os.path.join(self.source_folder, item)
                    if os.path.isfile(full_path):
                        source_files[item] = full_path
        except Exception as e:
            self.result_text.append(f"❌ 扫描源文件夹失败：{str(e)}\n")
            return

        # 检查哪些文件需要复制
        for rel_path, src_path in source_files.items():
            target_path = os.path.join(self.target_folder, rel_path)
            if not os.path.exists(target_path):
                files_to_copy.append(rel_path)
            elif self.check_overwrite.isChecked():
                src_size = os.path.getsize(src_path)
                dst_size = os.path.getsize(target_path)
                if src_size != dst_size:
                    files_to_copy.append(rel_path)

        # 检查需要删除的文件
        if self.check_delete.isChecked():
            try:
                if self.check_subfolders.isChecked():
                    for root, dirs, files in os.walk(self.target_folder):
                        for file in files:
                            full_path = os.path.join(root, file)
                            rel_path = os.path.relpath(
                                full_path, self.target_folder)
                            if rel_path not in source_files:
                                files_to_delete.append(rel_path)
                else:
                    for item in os.listdir(self.target_folder):
                        full_path = os.path.join(self.target_folder, item)
                        if os.path.isfile(full_path) and item not in source_files:
                            files_to_delete.append(item)
            except Exception as e:
                self.result_text.append(f"❌ 扫描目标文件夹失败：{str(e)}\n")

        # 显示预览结果
        self.result_text.append(f"📊 同步预览结果：\n")
        self.result_text.append(f"源文件夹: {self.source_folder}\n")
        self.result_text.append(f"目标文件夹: {self.target_folder}\n\n")

        if files_to_copy:
            self.result_text.append(
                f"📁 需要复制/更新的文件 ({len(files_to_copy)} 个):\n")
            for f in files_to_copy[:20]:
                self.result_text.append(f"   └─ {f}\n")
            if len(files_to_copy) > 20:
                self.result_text.append(
                    f"   ... 还有 {len(files_to_copy)-20} 个文件\n")
            self.result_text.append("\n")
        else:
            self.result_text.append("✅ 没有需要复制或更新的文件\n\n")

        if files_to_delete:
            self.result_text.append(
                f"🗑️ 需要删除的文件 ({len(files_to_delete)} 个):\n")
            for f in files_to_delete[:20]:
                self.result_text.append(f"   └─ {f}\n")
            if len(files_to_delete) > 20:
                self.result_text.append(
                    f"   ... 还有 {len(files_to_delete)-20} 个文件\n")
            self.result_text.append("\n")
        elif self.check_delete.isChecked():
            self.result_text.append("✅ 没有需要删除的文件\n\n")

        if not files_to_copy and not files_to_delete:
            self.result_text.append("🎉 文件夹已经同步完成，无需任何操作！\n")

    def start_sync(self):
        if not self.source_folder or not self.target_folder:
            QMessageBox.warning(self, "提示", "请先选择源文件夹和目标文件夹！")
            return

        # 确认对话框
        reply = QMessageBox.question(self, "确认同步",
                                     f"确定要将\n{self.source_folder}\n同步到\n{self.target_folder}\n吗？",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.btn_sync.setEnabled(False)
        self.btn_preview.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.result_text.clear()
        self.result_text.append("🔄 开始同步...\n")
        self.result_text.append("=" * 60 + "\n\n")

        # 收集源文件夹中的所有文件
        source_files = {}
        try:
            if self.check_subfolders.isChecked():
                for root, dirs, files in os.walk(self.source_folder):
                    for file in files:
                        full_path = os.path.join(root, file)
                        rel_path = os.path.relpath(
                            full_path, self.source_folder)
                        source_files[rel_path] = full_path
            else:
                for item in os.listdir(self.source_folder):
                    full_path = os.path.join(self.source_folder, item)
                    if os.path.isfile(full_path):
                        source_files[item] = full_path
        except Exception as e:
            self.result_text.append(f"❌ 扫描源文件夹失败：{str(e)}\n")
            self.progress.setVisible(False)
            self.btn_sync.setEnabled(True)
            self.btn_preview.setEnabled(True)
            return

        total_files = len(source_files)
        if total_files == 0:
            self.result_text.append("❌ 源文件夹中没有找到文件\n")
            self.progress.setVisible(False)
            self.btn_sync.setEnabled(True)
            self.btn_preview.setEnabled(True)
            return

        copied = 0
        failed = 0

        # 创建目标文件夹结构并复制文件
        import shutil

        for rel_path, src_path in source_files.items():
            target_path = os.path.join(self.target_folder, rel_path)
            target_dir = os.path.dirname(target_path)

            # 创建目标文件夹
            if not os.path.exists(target_dir):
                try:
                    os.makedirs(target_dir)
                except Exception as e:
                    self.result_text.append(
                        f"❌ 创建文件夹失败: {target_dir} - {str(e)}\n")
                    failed += 1
                    continue

            # 检查是否需要复制
            need_copy = False
            if not os.path.exists(target_path):
                need_copy = True
            elif self.check_overwrite.isChecked():
                src_size = os.path.getsize(src_path)
                dst_size = os.path.getsize(target_path)
                if src_size != dst_size:
                    need_copy = True

            if need_copy:
                try:
                    shutil.copy2(src_path, target_path)
                    copied += 1
                    self.result_text.append(f"✅ 复制: {rel_path}\n")
                except Exception as e:
                    failed += 1
                    self.result_text.append(f"❌ 复制失败: {rel_path} - {str(e)}\n")

            # 更新进度
            progress_value = int((copied + failed) / total_files * 80)
            self.progress.setValue(progress_value)
            QApplication.processEvents()

        # 删除多余的文件
        deleted = 0
        if self.check_delete.isChecked():
            self.result_text.append("\n🗑️ 正在删除多余文件...\n")

            try:
                if self.check_subfolders.isChecked():
                    for root, dirs, files in os.walk(self.target_folder):
                        for file in files:
                            full_path = os.path.join(root, file)
                            rel_path = os.path.relpath(
                                full_path, self.target_folder)
                            if rel_path not in source_files:
                                try:
                                    os.remove(full_path)
                                    deleted += 1
                                    self.result_text.append(
                                        f"🗑️ 删除: {rel_path}\n")
                                except Exception as e:
                                    self.result_text.append(
                                        f"❌ 删除失败: {rel_path} - {str(e)}\n")
                else:
                    for item in os.listdir(self.target_folder):
                        full_path = os.path.join(self.target_folder, item)
                        if os.path.isfile(full_path) and item not in source_files:
                            try:
                                os.remove(full_path)
                                deleted += 1
                                self.result_text.append(f"🗑️ 删除: {item}\n")
                            except Exception as e:
                                self.result_text.append(
                                    f"❌ 删除失败: {item} - {str(e)}\n")
            except Exception as e:
                self.result_text.append(f"❌ 删除过程出错：{str(e)}\n")

            self.progress.setValue(90)

        self.progress.setValue(100)

        # 显示完成信息
        self.result_text.append("\n" + "=" * 60 + "\n")
        self.result_text.append(f"✅ 同步完成！\n")
        self.result_text.append(f"📁 复制/更新文件: {copied} 个\n")
        if self.check_delete.isChecked():
            self.result_text.append(f"🗑️ 删除文件: {deleted} 个\n")
        if failed > 0:
            self.result_text.append(f"❌ 失败: {failed} 个\n")

        self.progress.setVisible(False)
        self.btn_sync.setEnabled(True)
        self.btn_preview.setEnabled(True)

        QMessageBox.information(self, "完成", "文件夹同步完成！")


# ========== 功能3：磁盘空间分析器 ==========
class DiskAnalyzerPage(QWidget):
    def __init__(self):
        super().__init__()
        self.current_path = None
        self.scan_thread = None
        # 配置中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei',
                                           'Microsoft YaHei', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

        self.setup_ui()

    def setup_ui(self):
        """初始化UI界面"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 标题
        title = QLabel("💾 磁盘空间分析器")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # 统一的 QGroupBox 字体
        group_font = QFont()
        group_font.setPointSize(13)

        # 路径选择区域
        path_group = QGroupBox("选择要分析的路径")
        path_group.setFont(group_font)  # 添加字体设置
        path_layout = QHBoxLayout()

        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("请选择要分析的文件夹或磁盘...")
        self.path_edit.setReadOnly(True)

        self.btn_browse = QPushButton("📂 选择文件夹")
        self.btn_browse.clicked.connect(self.select_path)
        self.btn_browse.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        path_layout.addWidget(self.path_edit)
        path_layout.addWidget(self.btn_browse)
        path_group.setLayout(path_layout)
        layout.addWidget(path_group)

        # 扫描选项
        option_group = QGroupBox("扫描选项")
        option_group.setFont(group_font)  # 添加字体设置
        option_layout = QVBoxLayout()

        self.check_subfolders = QCheckBox("包含子文件夹")
        check_font = QFont()
        check_font.setPointSize(10)
        self.check_subfolders.setFont(check_font)
        self.check_subfolders.setChecked(True)
        self.check_subfolders.setStyleSheet("""
            QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
            }
        """)

        self.check_large_files = QCheckBox("显示大文件列表 (大于100MB)")
        check_font = QFont()
        check_font.setPointSize(10)
        self.check_large_files.setFont(check_font)
        self.check_large_files.setChecked(True)
        self.check_large_files.setStyleSheet("""
            QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
            }
        """)

        option_layout.addWidget(self.check_subfolders)
        option_layout.addWidget(self.check_large_files)
        option_group.setLayout(option_layout)
        layout.addWidget(option_group)

        # 按钮区域
        button_layout = QHBoxLayout()

        self.btn_scan = QPushButton("🔍 开始扫描")
        self.btn_scan.clicked.connect(self.start_scan)
        self.btn_scan.setEnabled(False)
        self.btn_scan.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        self.btn_stop = QPushButton("⏹ 停止扫描")
        self.btn_stop.clicked.connect(self.stop_scan)
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet("""
            QPushButton {
                background-color: #FF6B6B;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF5252;
            }
        """)

        button_layout.addWidget(self.btn_scan)
        button_layout.addWidget(self.btn_stop)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # 结果显示区域
        result_layout = QHBoxLayout()

        # 左侧：饼图
        self.figure = plt.figure(figsize=(5, 5), dpi=100)
        self.canvas = FigureCanvas(self.figure)
        self.ax = self.figure.add_subplot(111)

        chart_group = QGroupBox("空间分布")
        chart_group.setFont(group_font)  # 添加字体设置
        chart_layout = QVBoxLayout()
        chart_layout.addWidget(self.canvas)
        chart_group.setLayout(chart_layout)

        # 右侧：文件列表
        list_group = QGroupBox("大文件列表")
        list_group.setFont(group_font)  # 添加字体设置
        list_layout = QVBoxLayout()

        self.file_list = QTreeWidget()
        self.file_list.itemDoubleClicked.connect(self.open_file_location)
        self.file_list.setHeaderLabels(["文件名", "大小", "路径", "删除建议"])
        self.file_list.setColumnWidth(0, 200)
        self.file_list.setColumnWidth(1, 100)
        self.file_list.setColumnWidth(2, 300)
        self.file_list.setColumnWidth(3, 100)

        list_layout.addWidget(self.file_list)
        list_group.setLayout(list_layout)

        result_layout.addWidget(chart_group, 1)
        result_layout.addWidget(list_group, 2)

        layout.addLayout(result_layout)

        # 状态栏
        self.status_label = QLabel("请选择要分析的路径")
        self.status_label.setStyleSheet("color: #666;")
        layout.addWidget(self.status_label)

    def select_path(self):
        """选择要分析的路径"""
        path = QFileDialog.getExistingDirectory(self, "选择要分析的文件夹")
        if path:
            self.current_path = path
            self.path_edit.setText(path)
            self.btn_scan.setEnabled(True)
            self.status_label.setText(f"已选择路径: {path}")

    def start_scan(self):
        """开始扫描"""
        if not self.current_path:
            QMessageBox.warning(self, "提示", "请先选择要分析的路径！")
            return

        # 清空之前的结果
        self.file_list.clear()
        self.ax.clear()
        self.canvas.draw()

        # 禁用扫描按钮，启用停止按钮
        self.btn_scan.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.progress.setVisible(True)
        self.progress.setValue(0)

        # 创建并启动扫描线程
        self.scan_thread = ScanThread(
            self.current_path,
            self.check_subfolders.isChecked(),
            self.check_large_files.isChecked()
        )
        self.scan_thread.progress_updated.connect(self.update_progress)
        self.scan_thread.scan_completed.connect(self.display_results)
        self.scan_thread.start()

    def stop_scan(self):
        """停止扫描"""
        if self.scan_thread and self.scan_thread.isRunning():
            self.scan_thread.stop()
            self.status_label.setText("扫描已停止")
            self.btn_scan.setEnabled(True)
            self.btn_stop.setEnabled(False)
            self.progress.setVisible(False)

    def update_progress(self, value, message):
        """更新进度"""
        self.progress.setValue(value)
        self.status_label.setText(message)

    def display_results(self, folder_sizes, large_files):
        """显示扫描结果"""
        # 绘制饼图
        if folder_sizes:
            self.plot_pie_chart(folder_sizes)

        # 显示大文件列表
        if large_files:
            self.display_large_files(large_files)

        # 更新状态
        self.status_label.setText("扫描完成！")
        self.btn_scan.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.progress.setVisible(False)

    def plot_pie_chart(self, folder_sizes):
        """绘制饼图"""
        # 获取前10个最大的文件夹
        top_folders = sorted(folder_sizes.items(),
                             key=lambda x: x[1], reverse=True)[:10]

        if not top_folders:
            return

        labels = [os.path.basename(path) for path, _ in top_folders]
        sizes = [size for _, size in top_folders]

        # 计算其他文件夹的总大小
        total_size = sum(folder_sizes.values())
        top_size = sum(sizes)
        if total_size > top_size:
            labels.append("其他")
            sizes.append(total_size - top_size)

        # 绘制饼图
        self.ax.clear()
        colors = plt.cm.Set3(range(len(labels)))
        wedges, texts, autotexts = self.ax.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%',
            startangle=90,
            colors=colors,
            wedgeprops={'linewidth': 1, 'edgecolor': 'white'}
        )

        # 设置文本样式
        for text in texts:
            text.set_fontsize(10)
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontsize(9)
            autotext.set_weight('bold')

        self.ax.set_title("文件夹空间分布", fontsize=12, pad=20)
        self.canvas.draw()

    # def display_large_files(self, large_files):
    #     """显示大文件列表"""
    #     self.file_list.clear()

    #     for file_path, file_size in large_files:
    #         # 格式化文件大小
    #         if file_size >= 1024 * 1024 * 1024:
    #             size_str = f"{file_size / (1024 * 1024 * 1024):.2f} GB"
    #         elif file_size >= 1024 * 1024:
    #             size_str = f"{file_size / (1024 * 1024):.2f} MB"
    #         else:
    #             size_str = f"{file_size / 1024:.2f} KB"

    #         # 创建树形项
    #         item = QTreeWidgetItem([
    #             os.path.basename(file_path),
    #             size_str,
    #             os.path.dirname(file_path)
    #         ])

    #         # 设置图标
    #         if os.path.isdir(file_path):
    #             item.setIcon(0, self.style().standardIcon(QStyle.SP_DirIcon))
    #         else:
    #             item.setIcon(0, self.style().standardIcon(QStyle.SP_FileIcon))

    #         self.file_list.addTopLevelItem(item)

    #     # 调整列宽
    #     self.file_list.resizeColumnToContents(0)
    #     self.file_list.resizeColumnToContents(1)
    def display_large_files(self, large_files):
        """显示大文件列表"""
        self.file_list.clear()

        for file_path, file_size, is_system in large_files:  # 解包时包含is_system标志
            # 格式化文件大小
            if file_size >= 1024 * 1024 * 1024:
                size_str = f"{file_size / (1024 * 1024 * 1024):.2f} GB"
            elif file_size >= 1024 * 1024:
                size_str = f"{file_size / (1024 * 1024):.2f} MB"
            else:
                size_str = f"{file_size / 1024:.2f} KB"

            # 判断是否为系统文件（使用扫描线程中已经判断好的标志）
            status = "不建议删除" if is_system else "可以考虑删除"

            # 创建树形项（包含4列：文件名、大小、路径、删除建议）
            item = QTreeWidgetItem([
                os.path.basename(file_path),
                size_str,
                os.path.dirname(file_path),
                status
            ])

            # 设置图标
            if os.path.isdir(file_path):
                item.setIcon(0, self.style().standardIcon(QStyle.SP_DirIcon))
            else:
                item.setIcon(0, self.style().standardIcon(QStyle.SP_FileIcon))

            # 设置状态颜色
            if is_system:
                item.setForeground(3, QBrush(Qt.red))
            else:
                item.setForeground(3, QBrush(Qt.green))

            self.file_list.addTopLevelItem(item)

        # 调整列宽
        self.file_list.resizeColumnToContents(0)
        self.file_list.resizeColumnToContents(1)
        self.file_list.resizeColumnToContents(3)

    # def is_system_file(self, file_path):
    #     """判断是否为系统文件"""
    #     # 获取文件名
    #     filename = os.path.basename(file_path).lower()

    #     # 系统文件扩展名列表
    #     system_extensions = [
    #         '.sys', '.dll', '.exe', '.com', '.bat', '.cmd',
    #         '.msi', '.msm', '.msp', '.mst', '.idb', '.pdb',
    #         '.lib', '.obj', '.res', '.manifest', '.config'
    #     ]

    #     # 系统文件夹列表
    #     system_folders = [
    #         'windows', 'program files', 'program files (x86)',
    #         'programdata', 'system32', 'syswow64'
    #     ]

    #     # 检查文件扩展名
    #     if any(filename.endswith(ext) for ext in system_extensions):
    #         return True

    #     # 检查文件路径是否包含系统文件夹
    #     path_lower = file_path.lower()
    #     if any(folder in path_lower for folder in system_folders):
    #         return True

    #     # 检查隐藏文件
    #     if os.name == 'nt':  # Windows系统
    #         try:
    #             import win32api
    #             import win32con
    #             attrs = win32api.GetFileAttributes(file_path)
    #             if attrs & (win32con.FILE_ATTRIBUTE_HIDDEN | win32con.FILE_ATTRIBUTE_SYSTEM):
    #                 return True
    #         except:
    #             pass

    #     return False

    def open_file_location(self, item):
        """打开文件所在位置"""
        # 获取完整文件路径
        file_name = item.text(0)  # 文件名
        file_dir = item.text(2)   # 文件所在目录

        # 确保路径格式正确
        if not file_dir:
            QMessageBox.warning(self, "错误", "无法获取文件路径")
            return

        # 拼接完整路径
        file_path = os.path.normpath(os.path.join(file_dir, file_name))

        # 验证文件是否存在
        if not os.path.exists(file_path):
            QMessageBox.warning(self, "错误", f"文件不存在：{file_path}")
            return

        try:
            if os.name == 'nt':  # Windows系统
                # 使用explorer的/select参数打开文件所在位置并选中文件
                subprocess.Popen(['explorer', '/select,', file_path])
            elif os.name == 'posix':  # Linux/Mac系统
                # 打开文件所在目录
                subprocess.Popen(['xdg-open', os.path.dirname(file_path)])
        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法打开文件位置：{str(e)}")


class ScanThread(QThread):
    """扫描文件夹的线程"""
    progress_updated = pyqtSignal(int, str)
    scan_completed = pyqtSignal(dict, list)

    def __init__(self, path, include_subfolders, show_large_files):
        super().__init__()
        self.path = path
        self.include_subfolders = include_subfolders
        self.show_large_files = show_large_files
        self._is_running = True

    def run(self):
        """执行扫描"""
        folder_sizes = {}
        large_files = []
        total_size = 0
        scanned_files = 0
        total_files = 0

        # 系统文件扩展名和文件夹列表（移到类成员变量）
        system_extensions = ['.sys', '.dll', '.exe', '.com', '.bat', '.cmd',
                             '.msi', '.msm', '.msp', '.mst', '.idb', '.pdb',
                             '.lib', '.obj', '.res', '.manifest', '.config']
        system_folders = ['windows', 'program files', 'program files (x86)',
                          'programdata', 'system32', 'syswow64']

        # 首先计算总文件数
        self.progress_updated.emit(5, "正在统计文件数量...")
        if self.include_subfolders:
            for root, dirs, files in os.walk(self.path):
                total_files += len(files)
        else:
            total_files = len([f for f in os.listdir(self.path)
                              if os.path.isfile(os.path.join(self.path, f))])

        if total_files == 0:
            self.progress_updated.emit(100, "没有找到文件")
            self.scan_completed.emit({}, [])
            return

        # 扫描文件
        self.progress_updated.emit(10, "正在扫描文件...")
        if self.include_subfolders:
            for root, dirs, files in os.walk(self.path):
                if not self._is_running:
                    break

                # 统计当前文件夹大小
                folder_size = 0
                for file in files:
                    if not self._is_running:
                        break

                    file_path = os.path.join(root, file)
                    try:
                        file_size = os.path.getsize(file_path)
                        folder_size += file_size
                        total_size += file_size

                        # 检查是否为大文件
                        if self.show_large_files and file_size > 100 * 1024 * 1024:
                            # 在扫描线程中判断是否为系统文件
                            is_system = self._is_system_file(
                                file_path, system_extensions, system_folders)
                            large_files.append(
                                (file_path, file_size, is_system))

                        scanned_files += 1
                        if scanned_files % 10 == 0:
                            progress = 10 + \
                                int(scanned_files / total_files * 80)
                            self.progress_updated.emit(
                                progress,
                                f"已扫描 {scanned_files}/{total_files} 个文件..."
                            )
                    except (OSError, PermissionError):
                        pass

                # 记录文件夹大小
                if folder_size > 0:
                    folder_sizes[root] = folder_size
        else:
            for item in os.listdir(self.path):
                if not self._is_running:
                    break

                item_path = os.path.join(self.path, item)
                try:
                    if os.path.isfile(item_path):
                        file_size = os.path.getsize(item_path)
                        total_size += file_size

                        # 检查是否为大文件
                        if self.show_large_files and file_size > 100 * 1024 * 1024:
                            large_files.append((item_path, file_size))

                        scanned_files += 1
                        if scanned_files % 10 == 0:
                            progress = 10 + \
                                int(scanned_files / total_files * 80)
                            self.progress_updated.emit(
                                progress,
                                f"已扫描 {scanned_files}/{total_files} 个文件..."
                            )
                    elif os.path.isdir(item_path):
                        # 统计子文件夹大小
                        folder_size = 0
                        for root, dirs, files in os.walk(item_path):
                            for file in files:
                                if not self._is_running:
                                    break

                                file_path = os.path.join(root, file)
                                try:
                                    file_size = os.path.getsize(file_path)
                                    folder_size += file_size
                                    total_size += file_size

                                    # 检查是否为大文件
                                    if self.show_large_files and file_size > 100 * 1024 * 1024:
                                        large_files.append(
                                            (file_path, file_size))

                                    scanned_files += 1
                                    if scanned_files % 10 == 0:
                                        progress = 10 + \
                                            int(scanned_files /
                                                total_files * 80)
                                        self.progress_updated.emit(
                                            progress,
                                            f"已扫描 {scanned_files}/{total_files} 个文件..."
                                        )
                                except (OSError, PermissionError):
                                    pass

                        if folder_size > 0:
                            folder_sizes[item_path] = folder_size
                except (OSError, PermissionError):
                    pass

        # 按大小排序大文件
        large_files.sort(key=lambda x: x[1], reverse=True)

        # 完成扫描
        self.progress_updated.emit(
            100, f"扫描完成！总大小: {self.format_size(total_size)}")
        self.scan_completed.emit(folder_sizes, large_files)

    def _is_system_file(self, file_path, system_extensions, system_folders):
        """判断是否为系统文件（辅助方法）"""
        filename = os.path.basename(file_path).lower()

        # 检查文件扩展名
        if any(filename.endswith(ext) for ext in system_extensions):
            return True

        # 检查文件路径是否包含系统文件夹
        path_lower = file_path.lower()
        if any(folder in path_lower for folder in system_folders):
            return True

        return False

    def stop(self):
        """停止扫描"""
        self._is_running = False
        self.wait()

    @staticmethod
    def format_size(size):
        """格式化文件大小"""
        if size >= 1024 * 1024 * 1024:
            return f"{size / (1024 * 1024 * 1024):.2f} GB"
        elif size >= 1024 * 1024:
            return f"{size / (1024 * 1024):.2f} MB"
        elif size >= 1024:
            return f"{size / 1024:.2f} KB"
        else:
            return f"{size} B"


# ========== 功能4：PDF批量处理工具 ==========
class PDFBatchPage(QWidget):
    def __init__(self):
        super().__init__()
        self.pdf_files = []  # 存储选中的PDF文件列表
        self.setup_ui()

    def setup_ui(self):
        """初始化UI界面"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 标题
        title = QLabel("📄 PDF 批量处理工具")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # 添加注释文本
        subtitle = QLabel("该功能暂未测试")
        subtitle.setFont(QFont("Arial", 12))  # 使用较小的字体
        subtitle.setStyleSheet("color: #888888;")  # 设置为灰色
        layout.addWidget(subtitle)

        # 统一的 QGroupBox 字体
        group_font = QFont()
        group_font.setPointSize(13)

        # 文件选择区域
        file_group = QGroupBox("选择PDF文件")
        file_group.setFont(group_font)
        file_layout = QVBoxLayout(file_group)

        # 文件列表
        self.file_list = QListWidget()
        self.file_list.setDragDropMode(
            QAbstractItemView.InternalMove)  # 允许拖拽排序
        file_layout.addWidget(self.file_list)

        # 文件操作按钮
        file_btn_layout = QHBoxLayout()

        self.btn_add = QPushButton("➕ 添加文件")
        self.btn_add.clicked.connect(self.add_files)
        self.btn_add.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        self.btn_remove = QPushButton("➖ 移除选中")
        self.btn_remove.clicked.connect(self.remove_files)
        self.btn_remove.setStyleSheet("""
            QPushButton {
                background-color: #FF6B6B;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF5252;
            }
        """)

        self.btn_clear = QPushButton("🗑️ 清空列表")
        self.btn_clear.clicked.connect(self.clear_files)
        self.btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #999;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #777;
            }
        """)

        file_btn_layout.addWidget(self.btn_add)
        file_btn_layout.addWidget(self.btn_remove)
        file_btn_layout.addWidget(self.btn_clear)
        file_layout.addLayout(file_btn_layout)

        layout.addWidget(file_group)

        # 功能选项卡
        self.tab_widget = QTabWidget()

        # 合并PDF标签页
        self.merge_tab = QWidget()
        self.setup_merge_tab()
        self.tab_widget.addTab(self.merge_tab, "📑 合并PDF")

        # 拆分PDF标签页
        self.split_tab = QWidget()
        self.setup_split_tab()
        self.tab_widget.addTab(self.split_tab, "✂️ 拆分PDF")

        # 添加水印标签页
        self.watermark_tab = QWidget()
        self.setup_watermark_tab()
        self.tab_widget.addTab(self.watermark_tab, "💧 添加水印")

        # 提取文字标签页
        self.extract_tab = QWidget()
        self.setup_extract_tab()
        self.tab_widget.addTab(self.extract_tab, "📝 提取文字")

        # 加密/解密标签页
        self.encrypt_tab = QWidget()
        self.setup_encrypt_tab()
        self.tab_widget.addTab(self.encrypt_tab, "🔐 加密/解密")

        layout.addWidget(self.tab_widget)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # 结果显示区域
        result_group = QGroupBox("处理结果")
        result_group.setFont(group_font)
        result_layout = QVBoxLayout(result_group)

        self.result_text = QTextEdit()
        text_font = QFont()
        text_font.setPointSize(12)
        self.result_text.setFont(text_font)
        self.result_text.setReadOnly(True)
        self.result_text.setMinimumHeight(200)
        result_layout.addWidget(self.result_text)

        layout.addWidget(result_group)

    def setup_merge_tab(self):
        """设置合并PDF标签页"""
        layout = QVBoxLayout(self.merge_tab)
        layout.setSpacing(15)

        # 说明
        info = QLabel("将多个PDF文件合并为一个文件，支持拖拽调整合并顺序")
        info.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(info)

        # 输出文件设置
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("输出文件名:"))

        self.merge_output = QLineEdit()
        self.merge_output.setPlaceholderText("合并后的文件名.pdf")
        output_layout.addWidget(self.merge_output)

        self.btn_merge_browse = QPushButton("📂 选择保存位置")
        self.btn_merge_browse.clicked.connect(self.browse_output_file)
        self.btn_merge_browse.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)
        output_layout.addWidget(self.btn_merge_browse)

        layout.addLayout(output_layout)

        # 合并按钮
        self.btn_merge = QPushButton("🔗 开始合并")
        self.btn_merge.clicked.connect(self.merge_pdfs)
        self.btn_merge.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                min-height: 35px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        layout.addWidget(self.btn_merge)

        layout.addStretch()

    def setup_split_tab(self):
        """设置拆分PDF标签页"""
        layout = QVBoxLayout(self.split_tab)
        layout.setSpacing(15)

        # 说明
        info = QLabel("从PDF文件中提取指定页面，生成新的PDF文件")
        info.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(info)

        # 页面范围设置
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("页面范围:"))

        self.split_range = QLineEdit()
        self.split_range.setPlaceholderText("例如: 1,3,5-10,15")
        range_layout.addWidget(self.split_range)

        layout.addLayout(range_layout)

        # 输出文件设置
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("输出文件名:"))

        self.split_output = QLineEdit()
        self.split_output.setPlaceholderText("拆分后的文件名.pdf")
        output_layout.addWidget(self.split_output)

        self.btn_split_browse = QPushButton("📂 选择保存位置")
        self.btn_split_browse.clicked.connect(self.browse_output_file)
        self.btn_split_browse.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)
        output_layout.addWidget(self.btn_split_browse)

        layout.addLayout(output_layout)

        # 拆分按钮
        self.btn_split = QPushButton("✂️ 开始拆分")
        self.btn_split.clicked.connect(self.split_pdf)
        self.btn_split.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                min-height: 35px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        layout.addWidget(self.btn_split)

        layout.addStretch()

    def setup_watermark_tab(self):
        """设置添加水印标签页"""
        layout = QVBoxLayout(self.watermark_tab)
        layout.setSpacing(15)

        # 说明
        info = QLabel("为PDF文件添加文本水印")
        info.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(info)

        # 水印文本设置
        text_layout = QHBoxLayout()
        text_layout.addWidget(QLabel("水印文本:"))

        self.watermark_text = QLineEdit()
        self.watermark_text.setPlaceholderText("例如: 机密文档")
        text_layout.addWidget(self.watermark_text)

        layout.addLayout(text_layout)

        # 水印设置
        settings_layout = QFormLayout()

        self.watermark_opacity = QSlider(Qt.Horizontal)
        self.watermark_opacity.setRange(10, 100)
        self.watermark_opacity.setValue(30)
        self.watermark_opacity.setTickPosition(QSlider.TicksBelow)
        self.watermark_opacity.setTickInterval(10)
        settings_layout.addRow("透明度:", self.watermark_opacity)

        self.watermark_rotation = QSlider(Qt.Horizontal)
        self.watermark_rotation.setRange(0, 360)
        self.watermark_rotation.setValue(45)
        self.watermark_rotation.setTickPosition(QSlider.TicksBelow)
        self.watermark_rotation.setTickInterval(45)
        settings_layout.addRow("旋转角度:", self.watermark_rotation)

        self.watermark_color = QPushButton("选择颜色")
        self.watermark_color.clicked.connect(self.choose_watermark_color)
        self.watermark_color.setStyleSheet(
            "background-color: #999999; color: white;")
        settings_layout.addRow("水印颜色:", self.watermark_color)

        layout.addLayout(settings_layout)

        # 输出文件设置
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("输出文件名:"))

        self.watermark_output = QLineEdit()
        self.watermark_output.setPlaceholderText("添加水印后的文件名.pdf")
        output_layout.addWidget(self.watermark_output)

        self.btn_watermark_browse = QPushButton("📂 选择保存位置")
        self.btn_watermark_browse.clicked.connect(self.browse_output_file)
        self.btn_watermark_browse.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)
        output_layout.addWidget(self.btn_watermark_browse)

        layout.addLayout(output_layout)

        # 添加水印按钮
        self.btn_watermark = QPushButton("💧 添加水印")
        self.btn_watermark.clicked.connect(self.add_watermark)
        self.btn_watermark.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                min-height: 35px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        layout.addWidget(self.btn_watermark)

        layout.addStretch()

        # 水印颜色
        self.watermark_rgb = (153, 153, 153)  # 默认灰色

    def setup_extract_tab(self):
        """设置提取文字标签页"""
        layout = QVBoxLayout(self.extract_tab)
        layout.setSpacing(15)

        # 说明
        info = QLabel("从PDF文件中提取文字内容")
        info.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(info)

        # 提取选项
        options_layout = QVBoxLayout()

        self.extract_all = QCheckBox("提取所有页面")
        self.extract_all.setChecked(True)
        options_layout.addWidget(self.extract_all)

        page_range_layout = QHBoxLayout()
        page_range_layout.addWidget(QLabel("指定页面范围:"))

        self.extract_range = QLineEdit()
        self.extract_range.setPlaceholderText("例如: 1,3,5-10,15")
        self.extract_range.setEnabled(False)
        page_range_layout.addWidget(self.extract_range)

        options_layout.addLayout(page_range_layout)

        self.extract_all.toggled.connect(self.extract_range.setEnabled)

        layout.addLayout(options_layout)

        # 输出文件设置
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("输出文件名:"))

        self.extract_output = QLineEdit()
        self.extract_output.setPlaceholderText("提取的文字保存为.txt文件")
        output_layout.addWidget(self.extract_output)

        self.btn_extract_browse = QPushButton("📂 选择保存位置")
        self.btn_extract_browse.clicked.connect(self.browse_output_file)
        self.btn_extract_browse.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)
        output_layout.addWidget(self.btn_extract_browse)

        layout.addLayout(output_layout)

        # 提取文字按钮
        self.btn_extract = QPushButton("📝 提取文字")
        self.btn_extract.clicked.connect(self.extract_text)
        self.btn_extract.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                min-height: 35px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        layout.addWidget(self.btn_extract)

        layout.addStretch()

    def setup_encrypt_tab(self):
        """设置加密/解密标签页"""
        layout = QVBoxLayout(self.encrypt_tab)
        layout.setSpacing(15)

        # 说明
        info = QLabel("为PDF文件添加密码保护或移除密码")
        info.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(info)

        # 操作选择
        operation_layout = QHBoxLayout()
        operation_layout.addWidget(QLabel("操作类型:"))

        self.encrypt_operation = QComboBox()
        self.encrypt_operation.addItems(["加密PDF", "解密PDF"])
        operation_layout.addWidget(self.encrypt_operation)

        layout.addLayout(operation_layout)

        # 密码设置
        password_layout = QHBoxLayout()
        password_layout.addWidget(QLabel("密码:"))

        self.encrypt_password = QLineEdit()
        self.encrypt_password.setEchoMode(QLineEdit.Password)
        password_layout.addWidget(self.encrypt_password)

        layout.addLayout(password_layout)

        # 输出文件设置
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("输出文件名:"))

        self.encrypt_output = QLineEdit()
        self.encrypt_output.setPlaceholderText("处理后的文件名.pdf")
        output_layout.addWidget(self.encrypt_output)

        self.btn_encrypt_browse = QPushButton("📂 选择保存位置")
        self.btn_encrypt_browse.clicked.connect(self.browse_output_file)
        self.btn_encrypt_browse.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)
        output_layout.addWidget(self.btn_encrypt_browse)

        layout.addLayout(output_layout)

        # 加密/解密按钮
        self.btn_encrypt = QPushButton("🔐 开始处理")
        self.btn_encrypt.clicked.connect(self.encrypt_pdf)
        self.btn_encrypt.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                min-height: 35px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        layout.addWidget(self.btn_encrypt)

        layout.addStretch()

    def add_files(self):
        """添加PDF文件到列表"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择PDF文件", "", "PDF文件 (*.pdf)")

        if files:
            for file in files:
                if file not in self.pdf_files:
                    self.pdf_files.append(file)
                    self.file_list.addItem(os.path.basename(file))

    def remove_files(self):
        """从列表中移除选中的文件"""
        selected_items = self.file_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "提示", "请先选择要移除的文件！")
            return

        for item in selected_items:
            row = self.file_list.row(item)
            self.file_list.takeItem(row)
            self.pdf_files.pop(row)

    def clear_files(self):
        """清空文件列表"""
        if not self.pdf_files:
            return

        reply = QMessageBox.question(
            self, "确认清空", "确定要清空所有文件吗？",
            QMessageBox.Yes | QMessageBox.No)

        if reply == QMessageBox.Yes:
            self.pdf_files.clear()
            self.file_list.clear()

    def browse_output_file(self):
        """浏览并选择输出文件位置"""
        sender = self.sender()
        if sender == self.btn_merge_browse:
            output = self.merge_output
            default_name = "合并后的文件.pdf"
        elif sender == self.btn_split_browse:
            output = self.split_output
            default_name = "拆分后的文件.pdf"
        elif sender == self.btn_watermark_browse:
            output = self.watermark_output
            default_name = "添加水印后的文件.pdf"
        elif sender == self.btn_extract_browse:
            output = self.extract_output
            default_name = "提取的文字.txt"
        elif sender == self.btn_encrypt_browse:
            output = self.encrypt_output
            default_name = "处理后的文件.pdf"
        else:
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "选择保存位置", default_name,
            "PDF文件 (*.pdf);;文本文件 (*.txt)")

        if file_path:
            output.setText(file_path)

    def choose_watermark_color(self):
        """选择水印颜色"""
        color = QColorDialog.getColor()
        if color.isValid():
            self.watermark_rgb = (color.red(), color.green(), color.blue())
            self.watermark_color.setStyleSheet(
                f"background-color: {color.name()}; color: white;")

    def merge_pdfs(self):
        """合并多个PDF文件"""
        if len(self.pdf_files) < 2:
            QMessageBox.warning(self, "提示", "请至少选择两个PDF文件进行合并！")
            return

        output_path = self.merge_output.text()
        if not output_path:
            QMessageBox.warning(self, "提示", "请指定输出文件名！")
            return

        self.result_text.clear()
        self.result_text.append("🔗 开始合并PDF文件...\n")
        self.result_text.append("=" * 60 + "\n\n")

        self.progress.setVisible(True)
        self.progress.setValue(0)

        try:
            merger = PyPDF2.PdfMerger()

            for i, pdf_file in enumerate(self.pdf_files):
                self.result_text.append(
                    f"正在处理: {os.path.basename(pdf_file)}\n")
                merger.append(pdf_file)

                # 更新进度
                progress = int((i + 1) / len(self.pdf_files) * 90)
                self.progress.setValue(progress)
                QApplication.processEvents()

            # 保存合并后的PDF
            merger.write(output_path)
            merger.close()

            self.progress.setValue(100)
            self.result_text.append(f"\n✅ 合并完成！文件已保存到: {output_path}\n")

            # 询问是否打开文件
            reply = QMessageBox.question(
                self, "完成", "PDF合并完成！是否打开文件？",
                QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                if os.name == 'nt':  # Windows系统
                    os.startfile(output_path)
                elif os.name == 'posix':  # Linux/Mac系统
                    subprocess.Popen(['xdg-open', output_path])

        except Exception as e:
            self.result_text.append(f"\n❌ 合并失败: {str(e)}\n")
            QMessageBox.critical(self, "错误", f"合并PDF失败: {str(e)}")

        self.progress.setVisible(False)

    def split_pdf(self):
        """拆分PDF文件"""
        if len(self.pdf_files) != 1:
            QMessageBox.warning(self, "提示", "请选择一个PDF文件进行拆分！")
            return

        output_path = self.split_output.text()
        if not output_path:
            QMessageBox.warning(self, "提示", "请指定输出文件名！")
            return

        page_range = self.split_range.text()
        if not page_range:
            QMessageBox.warning(self, "提示", "请指定要提取的页面范围！")
            return

        self.result_text.clear()
        self.result_text.append("✂️ 开始拆分PDF文件...\n")
        self.result_text.append("=" * 60 + "\n\n")

        self.progress.setVisible(True)
        self.progress.setValue(0)

        try:
            # 解析页面范围
            pages_to_extract = self.parse_page_range(page_range)
            if not pages_to_extract:
                QMessageBox.warning(self, "提示", "无效的页面范围！")
                self.progress.setVisible(False)
                return

            # 打开PDF文件
            pdf_reader = PyPDF2.PdfReader(self.pdf_files[0])
            total_pages = len(pdf_reader.pages)

            self.result_text.append(f"文件总页数: {total_pages}\n")
            self.result_text.append(
                f"提取页面: {', '.join(map(str, pages_to_extract))}\n\n")

            # 创建新的PDF
            pdf_writer = PyPDF2.PdfWriter()

            # 添加指定页面
            for page_num in pages_to_extract:
                if page_num < 1 or page_num > total_pages:
                    self.result_text.append(f"⚠️ 页面 {page_num} 超出范围，已跳过\n")
                    continue

                pdf_writer.add_page(pdf_reader.pages[page_num - 1])
                self.result_text.append(f"✅ 已添加页面 {page_num}\n")

                # 更新进度
                progress = int(pages_to_extract.index(
                    page_num) / len(pages_to_extract) * 90)
                self.progress.setValue(progress)
                QApplication.processEvents()

            # 保存拆分后的PDF
            with open(output_path, 'wb') as output_file:
                pdf_writer.write(output_file)

            self.progress.setValue(100)
            self.result_text.append(f"\n✅ 拆分完成！文件已保存到: {output_path}\n")

            # 询问是否打开文件
            reply = QMessageBox.question(
                self, "完成", "PDF拆分完成！是否打开文件？",
                QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                if os.name == 'nt':  # Windows系统
                    os.startfile(output_path)
                elif os.name == 'posix':  # Linux/Mac系统
                    subprocess.Popen(['xdg-open', output_path])

        except Exception as e:
            self.result_text.append(f"\n❌ 拆分失败: {str(e)}\n")
            QMessageBox.critical(self, "错误", f"拆分PDF失败: {str(e)}")

        self.progress.setVisible(False)

    def add_watermark(self):
        """为PDF添加水印"""
        if len(self.pdf_files) != 1:
            QMessageBox.warning(self, "提示", "请选择一个PDF文件添加水印！")
            return

        output_path = self.watermark_output.text()
        if not output_path:
            QMessageBox.warning(self, "提示", "请指定输出文件名！")
            return

        watermark_text = self.watermark_text.text()
        if not watermark_text:
            QMessageBox.warning(self, "提示", "请输入水印文本！")
            return

        self.result_text.clear()
        self.result_text.append("💧 开始添加水印...\n")
        self.result_text.append("=" * 60 + "\n\n")

        self.progress.setVisible(True)
        self.progress.setValue(0)

        try:
            # 创建水印PDF
            watermark_pdf = self.create_watermark_pdf(
                watermark_text,
                self.watermark_opacity.value() / 100,
                self.watermark_rotation.value(),
                self.watermark_rgb
            )

            # 打开原始PDF
            pdf_reader = PyPDF2.PdfReader(self.pdf_files[0])
            pdf_writer = PyPDF2.PdfWriter()

            # 为每一页添加水印
            total_pages = len(pdf_reader.pages)
            for i in range(total_pages):
                page = pdf_reader.pages[i]
                page.merge_page(watermark_pdf.pages[0])
                pdf_writer.add_page(page)

                # 更新进度
                progress = int((i + 1) / total_pages * 90)
                self.progress.setValue(progress)
                QApplication.processEvents()

            # 保存添加水印后的PDF
            with open(output_path, 'wb') as output_file:
                pdf_writer.write(output_file)

            self.progress.setValue(100)
            self.result_text.append(f"\n✅ 水印添加完成！文件已保存到: {output_path}\n")

            # 询问是否打开文件
            reply = QMessageBox.question(
                self, "完成", "水印添加完成！是否打开文件？",
                QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                if os.name == 'nt':  # Windows系统
                    os.startfile(output_path)
                elif os.name == 'posix':  # Linux/Mac系统
                    subprocess.Popen(['xdg-open', output_path])

        except Exception as e:
            self.result_text.append(f"\n❌ 添加水印失败: {str(e)}\n")
            QMessageBox.critical(self, "错误", f"添加水印失败: {str(e)}")

        self.progress.setVisible(False)

    def create_watermark_pdf(self, text, opacity, rotation, color):
        """创建水印PDF"""
        # 创建一个A4大小的PDF页面
        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=A4)

        # 设置水印样式
        can.setFillColorRGB(color[0]/255, color[1] /
                            255, color[2]/255, alpha=opacity)
        can.setFont("Helvetica", 60)

        # 保存当前状态
        can.saveState()

        # 旋转画布
        can.translate(A4[0] / 2, A4[1] / 2)
        can.rotate(rotation)

        # 绘制水印文本
        can.drawCentredString(0, 0, text)

        # 恢复状态
        can.restoreState()

        # 保存画布
        can.save()

        # 将BytesIO转换为PdfReader对象
        packet.seek(0)
        watermark_pdf = PyPDF2.PdfReader(packet)

        return watermark_pdf

    def extract_text(self):
        """从PDF中提取文字"""
        if len(self.pdf_files) != 1:
            QMessageBox.warning(self, "提示", "请选择一个PDF文件提取文字！")
            return

        output_path = self.extract_output.text()
        if not output_path:
            QMessageBox.warning(self, "提示", "请指定输出文件名！")
            return

        self.result_text.clear()
        self.result_text.append("📝 开始提取文字...\n")
        self.result_text.append("=" * 60 + "\n\n")

        self.progress.setVisible(True)
        self.progress.setValue(0)

        try:
            # 打开PDF文件
            with pdfplumber.open(self.pdf_files[0]) as pdf:
                total_pages = len(pdf.pages)

                # 确定要提取的页面
                if self.extract_all.isChecked():
                    pages_to_extract = range(1, total_pages + 1)
                else:
                    page_range = self.extract_range.text()
                    pages_to_extract = self.parse_page_range(page_range)
                    if not pages_to_extract:
                        QMessageBox.warning(self, "提示", "无效的页面范围！")
                        self.progress.setVisible(False)
                        return

                # 提取文字
                extracted_text = ""
                for i, page_num in enumerate(pages_to_extract):
                    if page_num < 1 or page_num > total_pages:
                        self.result_text.append(f"⚠️ 页面 {page_num} 超出范围，已跳过\n")
                        continue

                    page = pdf.pages[page_num - 1]
                    text = page.extract_text()
                    if text:
                        extracted_text += f"--- 第 {page_num} 页 ---\n\n{text}\n\n"

                    # 更新进度
                    progress = int((i + 1) / len(pages_to_extract) * 90)
                    self.progress.setValue(progress)
                    QApplication.processEvents()

                # 保存提取的文字
                with open(output_path, 'w', encoding='utf-8') as output_file:
                    output_file.write(extracted_text)

                self.progress.setValue(100)
                self.result_text.append(f"\n✅ 文字提取完成！已保存到: {output_path}\n")

                # 询问是否打开文件
                reply = QMessageBox.question(
                    self, "完成", "文字提取完成！是否打开文件？",
                    QMessageBox.Yes | QMessageBox.No)

                if reply == QMessageBox.Yes:
                    if os.name == 'nt':  # Windows系统
                        os.startfile(output_path)
                    elif os.name == 'posix':  # Linux/Mac系统
                        subprocess.Popen(['xdg-open', output_path])

        except Exception as e:
            self.result_text.append(f"\n❌ 提取文字失败: {str(e)}\n")
            QMessageBox.critical(self, "错误", f"提取文字失败: {str(e)}")

        self.progress.setVisible(False)

    def encrypt_pdf(self):
        """加密或解密PDF文件"""
        if len(self.pdf_files) != 1:
            QMessageBox.warning(self, "提示", "请选择一个PDF文件进行处理！")
            return

        output_path = self.encrypt_output.text()
        if not output_path:
            QMessageBox.warning(self, "提示", "请指定输出文件名！")
            return

        password = self.encrypt_password.text()
        if not password:
            QMessageBox.warning(self, "提示", "请输入密码！")
            return

        operation = self.encrypt_operation.currentText()

        self.result_text.clear()
        self.result_text.append(f"🔐 开始{operation}...\n")
        self.result_text.append("=" * 60 + "\n\n")

        self.progress.setVisible(True)
        self.progress.setValue(0)

        try:
            if operation == "加密PDF":
                # 加密PDF
                pdf_reader = PyPDF2.PdfReader(self.pdf_files[0])
                pdf_writer = PyPDF2.PdfWriter()

                # 复制所有页面
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)

                # 添加密码保护
                pdf_writer.encrypt(password)

                # 保存加密后的PDF
                with open(output_path, 'wb') as output_file:
                    pdf_writer.write(output_file)

                self.result_text.append(f"✅ PDF加密完成！密码已设置为: {password}\n")

            else:  # 解密PDF
                # 解密PDF
                pdf_reader = PyPDF2.PdfReader(self.pdf_files[0])

                # 检查PDF是否加密
                if not pdf_reader.is_encrypted:
                    QMessageBox.warning(self, "提示", "该PDF文件未加密！")
                    self.progress.setVisible(False)
                    return

                # 尝试解密
                try:
                    pdf_reader.decrypt(password)
                except:
                    QMessageBox.warning(self, "提示", "密码错误，无法解密PDF！")
                    self.progress.setVisible(False)
                    return

                pdf_writer = PyPDF2.PdfWriter()

                # 复制所有页面
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)

                # 保存解密后的PDF
                with open(output_path, 'wb') as output_file:
                    pdf_writer.write(output_file)

                self.result_text.append(f"✅ PDF解密完成！\n")

            self.progress.setValue(100)
            self.result_text.append(f"文件已保存到: {output_path}\n")

            # 询问是否打开文件
            reply = QMessageBox.question(
                self, "完成", f"{operation}完成！是否打开文件？",
                QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                if os.name == 'nt':  # Windows系统
                    os.startfile(output_path)
                elif os.name == 'posix':  # Linux/Mac系统
                    subprocess.Popen(['xdg-open', output_path])

        except Exception as e:
            self.result_text.append(f"\n❌ {operation}失败: {str(e)}\n")
            QMessageBox.critical(self, "错误", f"{operation}失败: {str(e)}")

        self.progress.setVisible(False)

    def parse_page_range(self, page_range):
        """解析页面范围字符串，返回页面列表"""
        pages = []
        parts = page_range.split(',')

        for part in parts:
            part = part.strip()
            if '-' in part:
                # 处理范围，如 "5-10"
                start, end = part.split('-')
                try:
                    start = int(start)
                    end = int(end)
                    pages.extend(range(start, end + 1))
                except ValueError:
                    continue
            else:
                # 处理单个页面，如 "5"
                try:
                    pages.append(int(part))
                except ValueError:
                    continue

        # 去重并排序
        pages = sorted(list(set(pages)))
        return pages


# ========== 功能5：Word/Excel 批量替换与生成 ==========
class OfficeBatchPage(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        label = QLabel("📝 Word/Excel 批量替换与生成")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        label.setFont(title_font)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("color: #666;")
        layout.addWidget(label)

        info = QLabel("此功能正在开发中，敬请期待...")
        info_font = QFont()
        info_font.setPointSize(14)
        info.setFont(info_font)
        info.setAlignment(Qt.AlignCenter)
        info.setStyleSheet("color: #999; margin-top: 20px;")
        layout.addWidget(info)


# ========== 功能6：文件名称提取器 ==========
# 新增功能: 文件名称提取器，批量提取文件名中的关键信息
class FileNameCut(QWidget):
    def __init__(self):
        super().__init__()
        self.folder_path = None
        self.extracted_data = []
        self.setup_ui()

    def setup_ui(self):
        """初始化UI界面"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 标题
        title = QLabel("✂️ 文件名称提取器")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # 文件夹选择区域
        folder_group = QGroupBox("选择文件夹")
        group_font = QFont()
        group_font.setPointSize(13)
        folder_group.setFont(group_font)
        folder_layout = QVBoxLayout(folder_group)

        # 文件夹路径选择行
        path_layout = QHBoxLayout()
        self.folder_path_edit = QLineEdit()
        self.folder_path_edit.setPlaceholderText("请选择包含要处理文件的文件夹...")
        self.folder_path_edit.setReadOnly(True)

        # 选择文件夹按钮
        btn_select = QPushButton("📂 选择文件夹")
        btn_select.setMinimumWidth(130)
        btn_select.clicked.connect(self.select_folder)
        btn_select.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        path_layout.addWidget(self.folder_path_edit)
        path_layout.addWidget(btn_select)
        folder_layout.addLayout(path_layout)

        # 包含子文件夹选项
        self.check_subfolders = QCheckBox("包含子文件夹")
        check_font = QFont()
        check_font.setPointSize(10)
        self.check_subfolders.setFont(check_font)
        self.check_subfolders.setChecked(False)
        self.check_subfolders.setStyleSheet("""
            QCheckBox {
                spacing: 5px;
            }
            QCheckBox::indicator {
                width: 30px;
                height: 30px;
            }
        """)
        folder_layout.addWidget(self.check_subfolders)

        layout.addWidget(folder_group)

        # 提取规则设置
        rule_group = QGroupBox("提取规则")
        rule_group.setFont(group_font)
        rule_layout = QVBoxLayout(rule_group)

        # 创建统一的水平布局
        settings_layout = QHBoxLayout()
        settings_layout.setSpacing(20)  # 增加间距

        # 分隔符设置
        separator_layout = QHBoxLayout()
        separator_label = QLabel("分隔符:")
        separator_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: #333;
                font-weight: bold;
                background-color: #f0f0f0;
                padding: 8px 12px;
                border-radius: 6px;
                min-width: 80px;
            }
        """)
        self.separator_edit = QLineEdit()
        # self.separator_edit.setPlaceholderText("_ 或 - 或 .")
        self.separator_edit.setText("_")
        self.separator_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #ddd;
                border-radius: 6px;
                min-height: 32px;
                font-size: 24px;
            }
        """)
        separator_layout.addWidget(separator_label)
        separator_layout.addWidget(self.separator_edit)
        settings_layout.addLayout(separator_layout)

        # 提取部分选择
        extract_layout = QHBoxLayout()
        extract_label = QLabel("提取部分:")
        extract_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: #333;
                font-weight: bold;
                background-color: #f0f0f0;
                padding: 8px 12px;
                border-radius: 6px;
                min-width: 80px;
            }
        """)
        self.extract_part = QSpinBox()
        self.extract_part.setMinimum(1)
        self.extract_part.setMaximum(10)
        self.extract_part.setValue(1)
        self.extract_part.setStyleSheet("""
            QSpinBox {
                padding: 8px 12px;
                border: 1px solid #ddd;
                border-radius: 6px;
                min-height: 32px;
                font-size: 24px;
            }
        """)
        extract_layout.addWidget(extract_label)
        extract_layout.addWidget(self.extract_part)
        settings_layout.addLayout(extract_layout)

        # 文件扩展名过滤
        filter_layout = QHBoxLayout()
        filter_label = QLabel("文件扩展名:")
        filter_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                color: #333;
                font-weight: bold;
                background-color: #f0f0f0;
                padding: 8px 12px;
                border-radius: 6px;
                min-width: 80px;
            }
        """)
        self.extension_edit = QLineEdit()
        self.extension_edit.setPlaceholderText(".pdf或.docx 不填默认所有文件类型")
        self.extension_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #ddd;
                border-radius: 6px;
                min-height: 32px;
                font-size: 24px;
            }
        """)
        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.extension_edit)
        settings_layout.addLayout(filter_layout)

        # 添加弹性空间，使控件靠左对齐
        settings_layout.addStretch()

        rule_layout.addLayout(settings_layout)
        layout.addWidget(rule_group)

        # 按钮区域
        button_layout = QHBoxLayout()

        self.btn_preview = QPushButton("👁️ 预览结果")
        self.btn_preview.clicked.connect(self.preview_extraction)
        self.btn_preview.setEnabled(False)
        self.btn_preview.setStyleSheet("""
            QPushButton {
                background-color: #FFB347;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FF9500;
            }
        """)

        self.btn_export = QPushButton("📊 导出到Excel")
        self.btn_export.clicked.connect(self.export_to_excel)
        self.btn_export.setEnabled(False)
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 6px;
                min-height: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)

        button_layout.addWidget(self.btn_preview)
        button_layout.addWidget(self.btn_export)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        # 结果显示区域
        result_group = QGroupBox("提取结果")
        result_group.setFont(group_font)
        result_layout = QVBoxLayout(result_group)

        self.result_table = QTableWidget()
        self.result_table.setColumnCount(2)
        self.result_table.setHorizontalHeaderLabels(["文件名", "提取内容"])
        self.result_table.horizontalHeader().setStretchLastSection(True)
        self.result_table.setMinimumHeight(300)
        # 确保表格行为正确
        self.result_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.NoEditTriggers)
        result_layout.addWidget(self.result_table)

        layout.addWidget(result_group)

        # 状态栏
        self.status_label = QLabel("请选择文件夹")
        self.status_label.setStyleSheet("color: #666;")
        layout.addWidget(self.status_label)

    def select_folder(self):
        """选择文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择文件夹")
        if folder:
            self.folder_path = folder
            self.folder_path_edit.setText(folder)
            self.btn_preview.setEnabled(True)
            self.status_label.setText(f"已选择文件夹: {folder}")

    def get_files(self):
        """获取要处理的文件列表"""
        files = []
        separator = self.separator_edit.text()
        extension = self.extension_edit.text().strip().lower()

        try:
            if self.check_subfolders.isChecked():
                # 包含子文件夹
                for root, dirs, filenames in os.walk(self.folder_path):
                    for filename in filenames:
                        # 检查文件扩展名
                        if not extension or filename.lower().endswith(extension):
                            files.append(os.path.join(root, filename))
                            # 限制文件数量
                            if len(files) >= 1000:
                                QMessageBox.warning(
                                    self, "提示", "文件数量过多，仅处理前1000个文件")
                                return files
            else:
                # 不包含子文件夹
                for filename in os.listdir(self.folder_path):
                    filepath = os.path.join(self.folder_path, filename)
                    if os.path.isfile(filepath):
                        # 检查文件扩展名
                        if not extension or filename.lower().endswith(extension):
                            files.append(filepath)
                            # 限制文件数量
                            if len(files) >= 1000:
                                QMessageBox.warning(
                                    self, "提示", "文件数量过多，仅处理前1000个文件")
                                return files
        except Exception as e:
            QMessageBox.warning(self, "错误", f"读取文件列表失败: {str(e)}")
            return []

        return files

    def preview_extraction(self):
        """预览提取结果"""
        try:
            if not self.folder_path:
                QMessageBox.warning(self, "提示", "请先选择文件夹！")
                return

            separator = self.separator_edit.text()
            if not separator:
                QMessageBox.warning(self, "提示", "请输入分隔符！")
                return

            part_index = self.extract_part.value() - 1  # 转换为0-based索引

            # 获取文件列表
            files = self.get_files()
            if not files:
                QMessageBox.warning(self, "提示", "没有找到符合条件的文件！")
                return

            # 清空之前的结果并确保列数正确
            self.result_table.clearContents()
            self.result_table.setRowCount(0)
            self.result_table.setColumnCount(2)
            self.extracted_data = []

            # 处理每个文件
            for filepath in files:
                filename = os.path.basename(filepath)
                # 根据分隔符分割文件名
                parts = filename.split(separator)

                # 提取指定部分
                if part_index < len(parts):
                    extracted = parts[part_index]
                else:
                    extracted = ""

                # 添加到表格
                row = self.result_table.rowCount()
                self.result_table.insertRow(row)
                self.result_table.setItem(row, 0, QTableWidgetItem(filename))
                self.result_table.setItem(row, 1, QTableWidgetItem(extracted))

                # 保存数据以便导出
                self.extracted_data.append((filename, extracted))

            # 更新状态
            self.status_label.setText(f"已处理 {len(files)} 个文件")
            self.btn_export.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"预览失败: {str(e)}")
            self.status_label.setText("预览失败")

    def export_to_excel(self):
        """导出结果到Excel"""
        if not self.extracted_data:
            QMessageBox.warning(self, "提示", "没有可导出的数据！")
            return

        # 选择保存位置
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", "提取结果.xlsx", "Excel文件 (*.xlsx)")

        if not file_path:
            return

        try:
            # 使用openpyxl创建Excel文件
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "文件名提取结果"

            # 设置表头
            headers = ["文件名", "提取内容"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # 填充数据
            for row, (filename, extracted) in enumerate(self.extracted_data, 2):
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value=extracted)

            # 调整列宽
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 30

            # 保存文件
            wb.save(file_path)

            # 更新状态
            self.status_label.setText(f"结果已导出到: {file_path}")

            # 询问是否打开文件
            reply = QMessageBox.question(
                self, "完成", "导出成功！是否打开文件？",
                QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                if os.name == 'nt':  # Windows系统
                    os.startfile(file_path)
                elif os.name == 'posix':  # Linux/Mac系统
                    subprocess.Popen(['xdg-open', file_path])

        except ImportError:
            QMessageBox.critical(
                self, "错误", "缺少openpyxl库，请先安装: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出Excel失败: {str(e)}")


class OCRPage(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        label = QLabel("🔍 OCR 文字识别工具")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        label.setFont(title_font)
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("color: #666;")
        layout.addWidget(label)

        info = QLabel("此功能正在开发中，敬请期待...")
        info_font = QFont()
        info_font.setPointSize(14)
        info.setFont(info_font)
        info.setAlignment(Qt.AlignCenter)
        info.setStyleSheet("color: #999; margin-top: 20px;")
        layout.addWidget(info)


def main():
    # 声明全局变量qApp，使其在整个模块中可用
    global qApp
    # 创建QApplication实例，sys.argv是命令行参数列表
    app = QApplication(sys.argv)
    # 将创建的QApplication实例赋值给全局变量qApp
    qApp = app
    # 创建主窗口实例
    window = PupAideMainWindow()
    # 显示主窗口
    window.show()
    # 启动应用程序的事件循环，并在应用程序退出时返回状态码
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
